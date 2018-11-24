from gurobipy import *
from openpyxl import *
import numpy as np

wb = load_workbook(filename='Tpp.xlsx', data_only=True)
ws = wb['Data']

nSources = ws['B1'].value
nSinks = ws['B2'].value

sources = range(nSources)
sinks = range(nSinks)

capacity = [ws['A' + str(i)].value for i in range(7, 11)]
demand = [ws[i + '6'].value for i in ['B', 'C', 'D', 'E', 'F']]

allCells = np.array([[cell.value for cell in row] for row in ws.iter_rows()])
cost = allCells[6:6 + nSources, 1:1 + nSinks]
fixedCost = allCells[6:6 + nSources, 10:10 + nSinks]

try:
    # Create a new model
    model = Model('Transportation Planning Problem')

    # Add variables
    qty = model.addVars(sources, sinks, vtype=GRB.INTEGER, lb=0, name="Quantity shipped")

    # Add constraints
    model.addConstrs((quicksum(qty[i, j] for j in sinks) <= capacity[i] for i in sources), name='Capacity constraint')
    model.addConstrs((quicksum(qty[i, j] for i in sources) == demand[j] for j in sinks), name='Demand constraint')

    # Objective function
    model.setObjective(quicksum(qty[i, j] * cost[i][j] for i in sources for j in sinks), GRB.MINIMIZE)

    model.optimize()

except GurobiError as e:
    print('Error:' + str(e))

try:
    binModel = Model('Transportation Planning wih fixed Cost')

    # Add variables
    qty = binModel.addVars(sources, sinks, vtype=GRB.INTEGER, lb=0, name="Quantity shipped")
    y = binModel.addVars(sources, sinks, vtype=GRB.BINARY, name="Incurs fix cost")

    # Add constraints
    binModel.addConstrs((quicksum(qty[i, j] for j in sinks) <= capacity[i] for i in sources),
                        name='Capacity constraint')
    binModel.addConstrs((quicksum(qty[i, j] for i in sources) == demand[j] for j in sinks), name='Demand constraint')
    binModel.addConstrs(qty[i, j] <= y[i, j] * max((max(capacity), max(demand))) for i in sources for j in sinks)

    for n in [0, 1, 4, 6]:
        # Objective function
        binModel.setObjective(quicksum(qty[i, j] * cost[i][j] for i in sources for j in sinks) +
                              quicksum(y[i, j] * fixedCost[i][j] * n for i in sources for j in sinks), GRB.MINIMIZE)

        binModel.optimize()


except GurobiError as e:
    print('Error:' + str(e))
